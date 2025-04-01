from openpyxl import Workbook, load_workbook
import random

def configs():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except: 
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="linha")
        config.cell(column=2, row=1, value=5)
        config.cell(column=1, row=2, value="coluna")
        config.cell(column=2, row=2, value=5)
        config.cell(column=1, row=3, value="dificuldade")
        config.cell(column=2, row=3, value=2)

    linhas      =  config.cell(column=2, row=1).value
    colunas     =  config.cell(column=2, row=2).value
    dificuldade =  config.cell(column=2, row=3).value

    wb.save('campo.xlsx')

    config = {
        'linha' : int(linhas),
        'coluna' : int(colunas),
        'dificuldade' : int(dificuldade)
    }

    return config

def criarTabuleiro():
    config = configs()
    quantLinha = config['linha']
    quantColuna = config['coluna']
    bombas = calculoBombas()
    tabuleiro = []
    cont = 1

    # 0 -> casa vazia | -1 -> bombas
    for i in range(int(quantLinha)):
        linha = []
        for j in range(0, int(quantColuna)):
            # print('{:5}'.format('0') , end='')
            if cont in bombas:
                linha.append(-1)
            else:
                linha.append(0)
            cont += 1
        # print()
        tabuleiro.append(linha)
    
    gravarTabuleiro(tabuleiro)

def gravarTabuleiro(tabuleiro):
    try:
        wb = load_workbook(filename='campo.xlsx')
    except:
        wb = Workbook()
    
    try:
        abaJogo = wb['jogo']
    except:
        abaJogo = wb.create_sheet('jogo')
    print()

    for linha in range(1, len(tabuleiro) +  1):
        for coluna in range(1, len(tabuleiro[0]) + 1):
            abaJogo.cell(row=linha, column=coluna, value=tabuleiro[linha - 1][coluna - 1])
    
    wb.save('campo.xlsx')
    print()

def configCampo():
    novaQuantColuna = input('Escolha a quantidade de colunas: ')
    novaQuantLinha = input('Escolha a quantidade de linhas: ')
    novaDificuldade = input('1- Facil \n2- Médio \n3- Dificil \nEscolha a dificuldade: ')

    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')

    config.cell(column=1, row=1, value="linha")
    config.cell(column=2, row=1, value=novaQuantLinha)
    config.cell(column=1, row=2, value="coluna")
    config.cell(column=2, row=2, value=novaQuantColuna)
    config.cell(column=1, row=3, value="dificuldade")
    config.cell(column=2, row=3, value=novaDificuldade)
    
    wb.save('campo.xlsx')

def calculoBombas():
    '''
    Descubra o valor total de 'casas' no tabuleiro e a partir desse total calcule a quantidade de bombas que vai ter no tabuleiro
    facil -> 15%, médio -> 30%, dificil -> 50% de bombas
    '''

    config = configs()
    totalCasas = config['linha'] * config['coluna']

    if config['dificuldade'] == 1:
        quantBomba = totalCasas * 0.15
    elif config['dificuldade'] == 2:
        quantBomba = totalCasas * 0.30
    elif config['dificuldade'] == 3:
        quantBomba = totalCasas * 0.50

    # print('total de bombas: {}'.format(int(quantBomba)))

    # Aleatorio bombas
    # [1, 2, 3], [4, 5, 6], [7, 8, 9]
    bombas = []
    while True:
        posicao = random.randint(1, totalCasas)
        if posicao not in bombas:
            bombas.append(posicao)
        if len(bombas) == int(quantBomba):
            break

    return bombas

def jogar():
    wb          = load_workbook(filename='campo.xlsx')
    jogo        = wb['jogo']
    maxLinha    = jogo.max_row
    maxColuna   = jogo.max_column
    gameOver    = False
    
    while True:
        linhaJogada  = int(input('Linha: '))
        colunaJogada = int(input('Coluna: '))
        # linhaJogada  = 3
        # colunaJogada = 2

        jogada = int(jogo.cell(row=linhaJogada, column=colunaJogada).value)
        if jogada == 0:
            jogo.cell(row=linhaJogada, column=colunaJogada, value=1)
        elif jogada == -1:
            jogo.cell(row=linhaJogada, column=colunaJogada, value=-2)
            gameOver = True
        elif jogada == 1:
            print('jogada já efetuada, tente novamente')

        wb.save('campo.xlsx')
        # 🤷, 👌, 💥
        for linha in range(1, maxLinha+1):
            for coluna in range(1, maxColuna+1):
                casa = jogo.cell(row=linha, column=coluna).value
                if (int(casa) == 0 or (int(casa) == -1 and gameOver == False)):
                    print('{:^3}'.format('[🤷]'), end='')
                elif int(casa) == 1:
                    print('{:^3}'.format('[😎]'), end='')
                elif int(casa) <= -1 and gameOver == True:
                    print('{:^3}'.format('[💥]'), end='')

            print()
        
        if gameOver == True:
            print('Perdeu, passa tudo 🏹')
            break

criarTabuleiro()
jogar()

# while True:
#     '''
#     Crie o menu de acesso do jogo com as seguintes opções "jogar, configurar, sair"
#     Hoje a gente vai criar o configurar, de as opções ao usuario para ele configurar o jogo dele
#     Nessas configurações coloque uma terceira config "dificuldade" o usuario vai poder escolher entre 1 - 3
#     '''
#     opcao = input('1- Jogar \n2- Configurar \n3- Sair \nEscolha uma das oções: ')
#     if opcao == '1':
#         print('Jogar')
#     elif opcao == '2':
#         print('Configurar')
#         configCampo()
#     elif opcao == '3':
#         print('Sair')
#         break
#     else:
#         print('\033[31mTu é besta?\033[0m')